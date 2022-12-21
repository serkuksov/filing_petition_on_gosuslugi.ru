import logging
from pprint import pprint
from openpyxl import load_workbook


def get_json_from_excel(path_exel: str) -> dict:
    """Создает json на основе данных всех листов Excel."""
    try:
        book = load_workbook(path_exel)
    except FileNotFoundError:
        logging.error('Не удалось найти файл Excel')
        raise
    data = {}
    for sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        row_generator = sheet.values
        try:
            headers = next(row_generator)
        except StopIteration:
            logging.error(f'Лист {sheet_name} книги пустой')
            raise Exception
        data[sheet_name] = []
        for row in row_generator:
            dict_row = dict(zip(headers, row))
            data[sheet_name].append(dict_row)
    book.close()
    return data


def get_not_filed_petitions(name_excel: str) -> list[dict]:
    """Получить список не поданых на госуслуги ходатайств
    с добавлением параметра номер строки"""
    petitions = get_json_from_excel(name_excel)['Лист1']
    not_filed_petitions = []
    i = 2
    for petition in petitions:
        if petition['Номер заявления'] is None and not petition['Номер ИП'] is None and not petition['Ходатайство'] is None:
            petition['number_rom'] = i
            not_filed_petitions.append(petition)
        i += 1
    if not_filed_petitions:
        logging.info(f'Получен список обращений требующих подачи {len(not_filed_petitions)} шт.')
    else:
        logging.info(f'Обращений требующих подачи нет')
    return not_filed_petitions


def save_filed_petition_in_excel(filed_petition: dict, name_excel: str):
    """Сохраняет информацию о ходатайстве json в Excel."""
    try:
        book = load_workbook(name_excel)
    except FileNotFoundError:
        logging.error('Не удалось найти файл Excel')
        raise
    list_excel = book['Лист1']
    row = filed_petition['number_rom']
    list_excel.cell(row=row, column=3).value = filed_petition['Дата отправки']
    list_excel.cell(row=row, column=4).value = filed_petition['Номер заявления']
    logging.info(f'Информация о подаче обращения № {filed_petition["Номер заявления"]} записана в Excel')
    book.save(name_excel)


if __name__ == '__main__':
    pprint(get_not_filed_petitions())
